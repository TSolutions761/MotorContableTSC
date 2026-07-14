from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from motor_contable_tsc.exportador_nubox import exportar_comprobantes_nubox
from motor_contable_tsc.models import (
    ComprobanteContable,
    DetalleAuxiliar,
    DetalleBancario,
    LineaContable,
)


def _crear_plantilla(ruta: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobantes"
    ws.append(
        [
            "Número",
            "Tipo",
            "Fecha",
            "Glosa",
            "Cuenta",
            "Debe",
            "Haber",
            "RUT",
            "Razón Social",
            "Monto Auxiliar",
            "Descripción Movimiento Bancario",
            "Número Documento Bancario",
            "Monto Bancario",
            "Fecha Bancaria",
        ]
    )
    ws.append([None] * 14)
    wb.save(ruta)


def test_exportar_comprobante_con_auxiliar_y_banco(tmp_path):
    plantilla = tmp_path / "plantilla.xlsx"
    salida = tmp_path / "salida.xlsx"
    _crear_plantilla(plantilla)

    comprobante = ComprobanteContable(
        numero=0,
        tipo="E",
        fecha=date(2026, 6, 30),
        glosa="PAGO PROVEEDORES",
        clave_grupo="202606-2101-01-C",
        lineas=[
            LineaContable(
                cuenta="2101-01",
                glosa="PAGO PROVEEDORES",
                debe=Decimal("150000"),
                haber=Decimal("0"),
                auxiliares=[
                    DetalleAuxiliar(
                        rut="18392974-7",
                        razon_social="PROVEEDOR DEMO SPA",
                        monto=Decimal("150000"),
                    )
                ],
            ),
            LineaContable(
                cuenta="1101-10",
                glosa="PAGO PROVEEDORES",
                debe=Decimal("0"),
                haber=Decimal("150000"),
                movimientos_bancarios=[
                    DetalleBancario(
                        descripcion="0183929747 TRANSFERENCIA",
                        numero_documento="0",
                        monto=Decimal("150000"),
                        fecha=date(2026, 6, 5),
                    )
                ],
            ),
        ],
    )

    resultado = exportar_comprobantes_nubox([comprobante], plantilla, salida)
    assert resultado == salida
    assert salida.exists()

    wb = load_workbook(salida, data_only=True)
    ws = wb["Comprobantes"]

    assert ws["A2"].value == 0
    assert ws["B2"].value == "E"
    assert ws["E2"].value == "2101-01"
    assert ws["F2"].value == 150000

    assert ws["H3"].value == "18392974-7"
    assert ws["I3"].value == "PROVEEDOR DEMO SPA"
    assert ws["J3"].value == 150000

    assert ws["E4"].value == "1101-10"
    assert ws["G4"].value == 150000

    assert ws["K5"].value == "0183929747 TRANSFERENCIA"
    assert ws["L5"].value == "0"
    assert ws["M5"].value == 150000
    assert ws["N5"].value == date(2026, 6, 5)


def test_rechaza_comprobante_descuadrado(tmp_path):
    plantilla = tmp_path / "plantilla.xlsx"
    salida = tmp_path / "salida.xlsx"
    _crear_plantilla(plantilla)

    comprobante = ComprobanteContable(
        numero=0,
        tipo="I",
        fecha=date(2026, 6, 30),
        glosa="INGRESO",
        clave_grupo="202606-1104-03-A",
        lineas=[
            LineaContable(cuenta="1101-10", glosa="INGRESO", debe=Decimal("100")),
            LineaContable(cuenta="1104-03", glosa="INGRESO", haber=Decimal("90")),
        ],
    )

    try:
        exportar_comprobantes_nubox([comprobante], plantilla, salida)
    except ValueError as exc:
        assert "descuadrado" in str(exc).lower()
    else:
        raise AssertionError("Se esperaba ValueError para comprobante descuadrado")
