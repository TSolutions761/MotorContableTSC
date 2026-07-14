from __future__ import annotations

import re
import unicodedata
from copy import copy
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import ComprobanteContable, LineaContable


_ALIAS_COLUMNAS: dict[str, tuple[str, ...]] = {
    "numero": ("NUMERO", "NÚMERO", "NRO COMPROBANTE", "NUMERO COMPROBANTE"),
    "tipo": ("TIPO", "TIPO COMPROBANTE"),
    "fecha": ("FECHA", "FECHA COMPROBANTE"),
    "glosa": ("GLOSA", "GLOSA COMPROBANTE"),
    "glosa_detalle": ("GLOSA DETALLE", "DETALLE", "GLOSA LINEA"),
    "cuenta": ("CUENTA", "CUENTA CONTABLE", "CODIGO CUENTA"),
    "debe": ("DEBE",),
    "haber": ("HABER",),
    "rut": ("RUT", "RUT AUXILIAR"),
    "razon_social": ("RAZON SOCIAL", "RAZÓN SOCIAL", "NOMBRE AUXILIAR"),
    "tipo_documento": ("TIPO DOCUMENTO", "TIPO DOC"),
    "folio": ("FOLIO", "NUMERO DOCUMENTO AUXILIAR", "NÚMERO DOCUMENTO AUXILIAR"),
    "monto_auxiliar": ("MONTO AUXILIAR", "MONTO DOCUMENTO", "MONTO AUX"),
    "fecha_auxiliar": (
        "FECHA AUXILIAR",
        "FECHA DOCUMENTO",
        "FECHA EMISION",
        "FECHA EMISIÓN",
        "FECHA VENCIMIENTO",
    ),
    "descripcion_bancaria": (
        "DESCRIPCION MOVIMIENTO BANCARIO",
        "DESCRIPCIÓN MOVIMIENTO BANCARIO",
        "DESCRIPCION MOVIMIENTO",
        "DESCRIPCIÓN MOVIMIENTO",
    ),
    "numero_documento_banco": (
        "NUMERO DOCUMENTO BANCARIO",
        "NÚMERO DOCUMENTO BANCARIO",
        "NUMERO DOCUMENTO BANCO",
        "NÚMERO DOCUMENTO BANCO",
        "NUMERO DOCUMENTO",
        "NÚMERO DOCUMENTO",
    ),
    "monto_banco": ("MONTO BANCARIO", "MONTO BANCO"),
    "fecha_banco": ("FECHA BANCARIA", "FECHA BANCO", "FECHA MOVIMIENTO"),
}


def _normalizar_texto(valor: object) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto.strip().upper())
    return texto


def _buscar_hoja_comprobantes(workbook):
    for nombre in workbook.sheetnames:
        if "COMPROB" in _normalizar_texto(nombre):
            return workbook[nombre]
    return workbook[workbook.sheetnames[0]]


def _detectar_fila_encabezados(ws, max_filas: int = 15) -> int:
    alias_normalizados = {
        _normalizar_texto(alias)
        for aliases in _ALIAS_COLUMNAS.values()
        for alias in aliases
    }
    mejor_fila = 0
    mejor_puntaje = 0
    for fila in range(1, min(ws.max_row, max_filas) + 1):
        valores = {_normalizar_texto(c.value) for c in ws[fila] if c.value is not None}
        puntaje = len(valores & alias_normalizados)
        if puntaje > mejor_puntaje:
            mejor_fila, mejor_puntaje = fila, puntaje
    if mejor_puntaje < 5:
        raise ValueError("No fue posible identificar la fila de encabezados de la plantilla Nubox")
    return mejor_fila


def _mapear_columnas(ws, fila_encabezados: int) -> dict[str, int]:
    encabezados = {
        _normalizar_texto(celda.value): celda.column
        for celda in ws[fila_encabezados]
        if celda.value is not None
    }
    resultado: dict[str, int] = {}
    for campo, aliases in _ALIAS_COLUMNAS.items():
        for alias in aliases:
            columna = encabezados.get(_normalizar_texto(alias))
            if columna is not None:
                resultado[campo] = columna
                break

    obligatorias = {"numero", "tipo", "fecha", "glosa", "cuenta", "debe", "haber"}
    faltantes = obligatorias - resultado.keys()
    if faltantes:
        raise ValueError(f"Faltan columnas obligatorias en la plantilla Nubox: {sorted(faltantes)}")
    return resultado


def _valor_decimal(valor: Decimal) -> int | float:
    if valor == valor.to_integral_value():
        return int(valor)
    return float(valor)


def _copiar_estilo_fila(ws, origen: int, destino: int) -> None:
    for col in range(1, ws.max_column + 1):
        origen_celda = ws.cell(origen, col)
        destino_celda = ws.cell(destino, col)
        if origen_celda.has_style:
            destino_celda._style = copy(origen_celda._style)
        if origen_celda.number_format:
            destino_celda.number_format = origen_celda.number_format
        if origen_celda.alignment:
            destino_celda.alignment = copy(origen_celda.alignment)
        if origen_celda.protection:
            destino_celda.protection = copy(origen_celda.protection)


def _escribir(ws, fila: int, columnas: dict[str, int], campo: str, valor: object) -> None:
    columna = columnas.get(campo)
    if columna is not None:
        ws.cell(fila, columna).value = valor


def _escribir_linea_contable(
    ws,
    fila: int,
    columnas: dict[str, int],
    linea: LineaContable,
    comprobante: ComprobanteContable,
    incluir_cabecera: bool,
) -> int:
    if incluir_cabecera:
        _escribir(ws, fila, columnas, "numero", comprobante.numero)
        _escribir(ws, fila, columnas, "tipo", comprobante.tipo)
        _escribir(ws, fila, columnas, "fecha", comprobante.fecha)
        _escribir(ws, fila, columnas, "glosa", comprobante.glosa)

    _escribir(ws, fila, columnas, "cuenta", linea.cuenta)
    _escribir(ws, fila, columnas, "glosa_detalle", linea.glosa)
    _escribir(ws, fila, columnas, "debe", _valor_decimal(linea.debe))
    _escribir(ws, fila, columnas, "haber", _valor_decimal(linea.haber))
    fila += 1

    for auxiliar in linea.auxiliares:
        _escribir(ws, fila, columnas, "rut", auxiliar.rut)
        _escribir(ws, fila, columnas, "razon_social", auxiliar.razon_social)
        _escribir(ws, fila, columnas, "monto_auxiliar", _valor_decimal(auxiliar.monto))
        fila += 1

    for detalle in linea.movimientos_bancarios:
        _escribir(ws, fila, columnas, "descripcion_bancaria", detalle.descripcion)
        _escribir(ws, fila, columnas, "numero_documento_banco", detalle.numero_documento or "0")
        _escribir(ws, fila, columnas, "monto_banco", _valor_decimal(detalle.monto))
        _escribir(ws, fila, columnas, "fecha_banco", detalle.fecha)
        fila += 1

    return fila


def exportar_comprobantes_nubox(
    comprobantes: Iterable[ComprobanteContable],
    plantilla: str | Path,
    salida: str | Path,
    *,
    eliminar_ejemplos: bool = True,
) -> Path:
    """Escribe comprobantes jerárquicos sobre la plantilla oficial Nubox.

    Cada comprobante se exporta como:
    1. línea contable con cabecera;
    2. detalles auxiliares de esa línea, si existen;
    3. segunda línea contable;
    4. detalles bancarios de esa línea.
    """
    ruta_plantilla = Path(plantilla)
    ruta_salida = Path(salida)
    workbook = load_workbook(ruta_plantilla)
    ws = _buscar_hoja_comprobantes(workbook)
    fila_encabezados = _detectar_fila_encabezados(ws)
    columnas = _mapear_columnas(ws, fila_encabezados)
    primera_fila_datos = fila_encabezados + 1

    if eliminar_ejemplos and ws.max_row >= primera_fila_datos:
        ws.delete_rows(primera_fila_datos, ws.max_row - primera_fila_datos + 1)

    fila = primera_fila_datos
    fila_estilo = primera_fila_datos
    for comprobante in comprobantes:
        if not comprobante.cuadrado:
            raise ValueError(f"No se puede exportar comprobante descuadrado: {comprobante.clave_grupo}")
        if len(comprobante.lineas) != 2:
            raise ValueError(
                f"El comprobante {comprobante.clave_grupo} debe tener exactamente dos líneas contables"
            )

        for indice, linea in enumerate(comprobante.lineas):
            _copiar_estilo_fila(ws, fila_estilo, fila)
            fila = _escribir_linea_contable(
                ws,
                fila,
                columnas,
                linea,
                comprobante,
                incluir_cabecera=indice == 0,
            )

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(ruta_salida)
    return ruta_salida
