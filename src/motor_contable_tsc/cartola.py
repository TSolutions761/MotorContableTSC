from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import GrupoContable, MovimientoBancario


_ALIAS_COLUMNAS = {
    "monto": ("MONTO",),
    "descripcion": (
        "DESCRIPCIÓN MOVIMIENTO - GLOSA CUENTA BANCARIA",
        "DESCRIPCION MOVIMIENTO - GLOSA CUENTA BANCARIA",
        "DESCRIPCIÓN MOVIMIENTO",
        "DESCRIPCION MOVIMIENTO",
    ),
    "observacion": ("OBSERVACION - GLOSA", "OBSERVACIÓN - GLOSA", "OBSERVACION"),
    "fecha": ("FECHA",),
    "cuenta": ("CUENTA CONTABLE",),
    "documento": ("NUMERO DE DOCUMENTO", "NÚMERO DE DOCUMENTO", "NUEMRO DE DOCUMENTO"),
    "cargo_abono": ("CARGO/ABONO",),
}


def _normalizar_encabezado(valor: object) -> str:
    return " ".join(str(valor or "").strip().upper().split())


def _buscar_hoja_cartola(workbook) -> object:
    for nombre in workbook.sheetnames:
        if "CARTOLA" in nombre.upper():
            return workbook[nombre]
    return workbook[workbook.sheetnames[0]]


def _mapear_columnas(ws) -> dict[str, int]:
    encabezados = {
        _normalizar_encabezado(celda.value): celda.column
        for celda in ws[1]
        if celda.value is not None
    }
    resultado: dict[str, int] = {}
    for campo, aliases in _ALIAS_COLUMNAS.items():
        for alias in aliases:
            columna = encabezados.get(_normalizar_encabezado(alias))
            if columna is not None:
                resultado[campo] = columna
                break
    faltantes = {"monto", "fecha", "cuenta", "cargo_abono"} - resultado.keys()
    if faltantes:
        raise ValueError(f"Faltan columnas obligatorias en la cartola: {sorted(faltantes)}")
    return resultado


def _convertir_fecha(valor: object) -> date:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor or "").strip()
    for formato in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha no reconocida: {valor!r}")


def _convertir_monto(valor: object) -> Decimal:
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))
    texto = str(valor or "").strip().replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ValueError(f"Monto no reconocido: {valor!r}") from exc


def leer_cartola(ruta: str | Path) -> list[MovimientoBancario]:
    """Lee la hoja CARTOLA y devuelve movimientos normalizados."""
    workbook = load_workbook(Path(ruta), data_only=True)
    ws = _buscar_hoja_cartola(workbook)
    columnas = _mapear_columnas(ws)
    movimientos: list[MovimientoBancario] = []

    for fila in range(2, ws.max_row + 1):
        fecha_valor = ws.cell(fila, columnas["fecha"]).value
        monto_valor = ws.cell(fila, columnas["monto"]).value
        if fecha_valor in (None, "") and monto_valor in (None, ""):
            continue

        cargo_abono = str(ws.cell(fila, columnas["cargo_abono"]).value or "").strip().upper()
        if cargo_abono not in {"A", "C"}:
            raise ValueError(f"Fila {fila}: CARGO/ABONO debe ser A o C; recibido {cargo_abono!r}")

        cuenta = str(ws.cell(fila, columnas["cuenta"]).value or "").strip()
        if not cuenta:
            raise ValueError(f"Fila {fila}: falta CUENTA CONTABLE")

        movimientos.append(
            MovimientoBancario(
                fecha=_convertir_fecha(fecha_valor),
                cargo_abono=cargo_abono,
                cuenta_contable=cuenta,
                monto=_convertir_monto(monto_valor),
                descripcion_bancaria=str(
                    ws.cell(fila, columnas.get("descripcion", 0)).value or ""
                ).strip()
                if columnas.get("descripcion")
                else "",
                observacion=str(
                    ws.cell(fila, columnas.get("observacion", 0)).value or ""
                ).strip()
                if columnas.get("observacion")
                else "",
                numero_documento=str(
                    ws.cell(fila, columnas.get("documento", 0)).value or "0"
                ).strip()
                if columnas.get("documento")
                else "0",
            )
        )
    return movimientos


def agrupar_movimientos(
    movimientos: Iterable[MovimientoBancario],
) -> list[GrupoContable]:
    """Agrupa por Mes + Cuenta Contable + Cargo/Abono, sin mezclar ingresos y egresos."""
    grupos: dict[tuple[str, str, str], list[MovimientoBancario]] = defaultdict(list)
    for movimiento in movimientos:
        grupos[movimiento.clave_grupo].append(movimiento)

    resultado = [
        GrupoContable(
            periodo=periodo,
            cuenta_contable=cuenta,
            cargo_abono=cargo_abono,
            movimientos=items,
        )
        for (periodo, cuenta, cargo_abono), items in grupos.items()
    ]
    return sorted(resultado, key=lambda g: (g.periodo, g.cuenta_contable, g.cargo_abono))
