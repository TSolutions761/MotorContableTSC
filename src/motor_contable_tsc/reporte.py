from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from .normalizador import MovimientoNormalizado


def exportar_normalizacion_excel(
    movimientos: Iterable[MovimientoNormalizado],
    ruta_salida: str | Path,
) -> Path:
    """Genera un Excel verificable con movimientos normalizados y resumen de grupos."""
    items = list(movimientos)
    destino = Path(ruta_salida)
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "MOVIMIENTOS_NORMALIZADOS"
    encabezados = [
        "PERIODO",
        "FECHA",
        "CARGO/ABONO",
        "TIPO NUBOX",
        "CUENTA CONTABLE",
        "MONTO",
        "RUT",
        "RAZON SOCIAL",
        "GLOSA CONTABLE",
        "GLOSA BANCARIA",
        "NUMERO DOCUMENTO",
        "CLAVE GRUPO",
        "ESTADO",
    ]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    grupos: dict[tuple[str, str, str], dict[str, object]] = defaultdict(
        lambda: {"cantidad": 0, "total": 0, "estados": set()}
    )

    for item in items:
        mov = item.movimiento
        clave_texto = f"{mov.periodo}-{mov.cuenta_contable}-{mov.cargo_abono}"
        ws.append(
            [
                mov.periodo,
                mov.fecha,
                mov.cargo_abono,
                mov.tipo_nubox,
                mov.cuenta_contable,
                float(mov.monto),
                item.rut,
                item.razon_social,
                item.glosa_contable,
                mov.descripcion_bancaria,
                mov.numero_documento,
                clave_texto,
                item.estado,
            ]
        )
        grupo = grupos[item.clave_grupo]
        grupo["cantidad"] = int(grupo["cantidad"]) + 1
        grupo["total"] = float(grupo["total"]) + float(mov.monto)
        cast_estados = grupo["estados"]
        assert isinstance(cast_estados, set)
        cast_estados.add(item.estado)

    resumen = wb.create_sheet("RESUMEN_GRUPOS")
    resumen.append(
        [
            "PERIODO",
            "CUENTA CONTABLE",
            "CARGO/ABONO",
            "TIPO NUBOX",
            "CANTIDAD MOVIMIENTOS",
            "MONTO TOTAL",
            "ESTADO",
        ]
    )
    for celda in resumen[1]:
        celda.font = Font(bold=True)

    for (periodo, cuenta, cargo_abono), datos in sorted(grupos.items()):
        estados = datos["estados"]
        assert isinstance(estados, set)
        resumen.append(
            [
                periodo,
                cuenta,
                cargo_abono,
                "I" if cargo_abono == "A" else "E",
                datos["cantidad"],
                datos["total"],
                "OK" if estados == {"OK"} else "REVISAR",
            ]
        )

    for hoja in (ws, resumen):
        hoja.freeze_panes = "A2"
        hoja.auto_filter.ref = hoja.dimensions

    wb.save(destino)
    return destino
